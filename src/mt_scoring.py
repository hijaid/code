# coding=utf8
import os
import pandas as pd
from tqdm import tqdm
import subprocess
import json
import shutil
from collections import defaultdict
import argparse
import datetime
from openpyxl import load_workbook,Workbook
from openpyxl.utils import get_column_letter

from sacrebleu.metrics import BLEU, CHRF, TER
from comet import  load_from_checkpoint


def bleu_scoring(ref_file, hypo_file, lp):
    src, tgt = lp.split("2")
    langpair = f"{src}-{tgt}"
    command = f"sacrebleu -w 2 -b {ref_file} -i {hypo_file} -l {langpair}"
    print(command)
    score = subprocess.run(command, stdout=subprocess.PIPE, stderr=subprocess.PIPE, shell=True, text=True)
    return float(score.stdout.strip()) 


def comet22_scoring(src_file, ref_file, hypo_file, model):
    srcs = [x.strip() for x in  open(src_file, encoding='utf-8')]
    refs = [x.strip() for x in  open(ref_file, encoding='utf-8')]
    hypos = [x.strip() for x in  open(hypo_file, encoding='utf-8')] 
    assert len(srcs) == len(refs) == len(hypos), print(src_file, ref_file, hypo_file)
    data = [{"src":x, "mt":y, "ref":z} for x,y,z in zip(srcs, hypos, refs)]
    print(f"comet22\nsrc_file: {src_file}\nref_file: {ref_file}\nhypo_file: {hypo_file}")
    model_output = model.predict(data, batch_size=128, gpus=1) ###256
    score = round(model_output[1]*100, 2)
    return score

def xcomet_scoring(src_file, hypo_file, model):
    srcs = [x.strip() for x in  open(src_file, encoding='utf-8') if x.strip()]
    hypos = [x.strip() for x in  open(hypo_file, encoding='utf-8') if x.strip()] 
    assert len(srcs) == len(hypos)
    data = [{"src":x, "mt":y} for x,y in zip(srcs, hypos)]
    print(f"xcomet\nsrc_file: {src_file}\nhypo_file: {hypo_file}")
    model_output = model.predict(data, batch_size=16, gpus=1)
    score = round(model_output[1]*100, 2)
    return score

def write_xlsl(file, data, flag=""):
    if os.path.exists(file):
        wb =  load_workbook(file)
    else:
        wb = Workbook()

    ws = wb.active

    # 找到第一个空白行的位置
    row_index = 1
    while ws[f'A{row_index}'].value is not None:
        row_index += 1
    
    current_time = datetime.datetime.now()
    ws[f'A{row_index}'] = f"{current_time.strftime('%Y-%m-%d %H:%M:%S')}\n{flag}"
    # ws[f'B{row_index}'] = flag

    headers = list(data.keys())
    for col_index, header in enumerate(headers, start=1):
        ws[f'{get_column_letter(col_index)}{row_index + 1}'] = header
    
    max_length = max(len(value) for value in data.values())
    for i in range(max_length):
        row_index += 1
        for col_index, (key, values) in enumerate(data.items(), start=1):
            try:
                ws[f'{get_column_letter(col_index)}{row_index + 1}'] = values[i]
            except:
                print(data)
                print(flag)
                print(values, max_length)

    wb.save(file)

def sort_data(src_files, hypo_files, ref_files, lang_pairs):
    sort_order = {'de2en': 1, 'cs2en': 2, 'ru2en': 3, 'zh2en': 4, 'en2de': 5,'en2cs': 6,'en2ru': 7,'en2zh': 8}
    combined = list(zip(src_files, hypo_files, ref_files, lang_pairs))
    combined_sorted = sorted(combined, key=lambda x: sort_order.get(x[-1], 100))
    src_files, hypo_files, ref_files, lang_pairs = zip(*combined_sorted)
    return list(src_files), list(hypo_files), list(ref_files), list(lang_pairs)


def main():
    parser = argparse.ArgumentParser(description="Script with conditional parameters")
    parser.add_argument('--metric', type=str, help='The evaluate metric', default="bleu,comet_22,xcomet_xxl")
    parser.add_argument('--comet_22_path', default="/mnt/luoyingfeng/model_card/wmt22-comet-da/checkpoints/model.ckpt", type=str, help='The comet22 path model')
    parser.add_argument('--xcomet_xl_path', default="/mnt/luoyingfeng/model_card/XCOMET-XL/checkpoints/model.ckpt", type=str, help='The xcomet xl path model')
    parser.add_argument('--xcomet_xxl_path', default="/mnt/luoyingfeng/model_card/XCOMET-XXL/checkpoints/model.ckpt", type=str, help='The xcomet xxl path model')
    parser.add_argument('--lang_pair', type=str, help='plain text')
    parser.add_argument('--write_key', type=str, default="language", help='plain text')
    parser.add_argument('--src_file', type=str, help='plain text')
    parser.add_argument('--ref_file', type=str, help='plain text')
    parser.add_argument('--hypo_file', type=str, help='plain text')
    parser.add_argument('--record_file', default="result.xlsx", type=str, help='plain text')
    parser.add_argument('--gpu', type=str, default="0,1,2,3,4,5", help='plain text')
    args = parser.parse_args()

    os.environ['CUDA_VISIBLE_DEVICES'] = args.gpu

    src_files = args.src_file.split(",")
    hypo_files = args.hypo_file.split(",")
    ref_files =  args.ref_file.split(",") 
    lang_pairs = args.lang_pair.split(",")
    assert len(src_files) == len(hypo_files) == len(lang_pairs) == len(ref_files)

    src_files, hypo_files, ref_files, lang_pairs = sort_data(src_files, hypo_files, ref_files, lang_pairs)
    metrics = args.metric.split(",")

    if "comet_22" in metrics:
        comet_22_model = load_from_checkpoint(args.comet_22_path)
    if "xcomet_xl" in metrics:
        comet_xl_model = load_from_checkpoint(args.xcomet_xl_path)
    if "xcomet_xxl" in metrics:
        comet_xxl_model = load_from_checkpoint(args.xcomet_xxl_path)
    
    result = defaultdict(list)
    result["metric"] = metrics
    for metric in metrics:
        for lp,src_file,ref_file, hypo_file in zip(lang_pairs, src_files, ref_files, hypo_files):
            if not os.path.isfile(src_file):
                print(f"file {src_file} not exist!")
                exit()
            if not os.path.isfile(ref_file):
                print(f"file {ref_file} not exist!")
                exit()
            print(f"evaluate {lp}")

            if args.write_key == "language":
                wk = lp
            else:
                # hypo suffix
                wk = os.path.basename(hypo_file)

            if metric == "bleu":
                score = bleu_scoring(ref_file, hypo_file, lp)
                result[wk].append(score)        
            
            if metric == "comet_22":
                score = comet22_scoring(src_file, ref_file, hypo_file, comet_22_model)
                result[wk].append(score)
            
            if metric == "xcomet_xl":
                score = xcomet_scoring(src_file, hypo_file, comet_xl_model)
                result[wk].append(score)
            
            if metric == "xcomet_xxl":
                score = xcomet_scoring(src_file, hypo_file, comet_xxl_model)
                result[wk].append(score)
    write_xlsl(args.record_file, result, flag=hypo_files[-1])


if __name__ == '__main__':
    main()

